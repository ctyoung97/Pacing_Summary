import pandas as pd
import openpyxl
import os
import re
from pathlib import Path

def get_station_worksheets(workbook):
    """
    Get all worksheet names that appear before 'MTH Summary' sheet.
    These are considered station worksheets.
    """
    sheet_names = workbook.sheetnames
    try:
        mth_summary_index = sheet_names.index('MTH Summary')
        station_sheets = sheet_names[:mth_summary_index]
        print(f"Found {len(station_sheets)} station worksheets: {station_sheets}")
        return station_sheets
    except ValueError:
        print("Warning: 'MTH Summary' sheet not found. Using all sheets.")
        return sheet_names

def extract_cell_data(worksheet, cell_address):
    """
    Safely extract data from a specific cell, returning 0 if cell is empty or contains error.
    """
    try:
        cell_value = worksheet[cell_address].value
        if cell_value is None:
            return 0
        # Handle potential string values or errors
        if isinstance(cell_value, str):
            return 0
        return float(cell_value)
    except Exception as e:
        print(f"Warning: Could not extract data from cell {cell_address}: {e}")
        return 0

def generate_q3_pacing_setup(current_file_path, prior_file_path, last_week_pace_data=None):
    """
    Generate Q3 Pacing Data Setup with Core Revenue report.
    Now includes New Core Billing data between columns B and C.
    Includes Last Week Q3 Pace and Pace Delta (bps) columns.
    Filters out unwanted rows and moves QTR Summary to bottom.
    """
    print(f"Processing Q3 Pacing Setup from: {current_file_path}")
    
    # Load both workbooks for the new billing calculation
    current_wb = openpyxl.load_workbook(current_file_path, data_only=True)
    prior_wb = openpyxl.load_workbook(prior_file_path, data_only=True)
    station_sheets = get_station_worksheets(current_wb)
    
    # Define rows to exclude
    excluded_rows = {'->', 'QTD Summ by Station', 'YTD Summ by Station', 
                     'Station Summary', 'Instructions', 'AMB Corp', '<-'}
    
    # Filter station sheets and separate QTR Summary
    filtered_sheets = []
    qtr_summary_data = None
    
    for sheet_name in station_sheets:
        if sheet_name in excluded_rows:
            print(f"Excluding sheet: {sheet_name}")
            continue
        elif sheet_name == 'QTR Summary':
            # Process QTR Summary but don't add to main list yet
            qtr_summary_data = sheet_name
            continue
        else:
            filtered_sheets.append(sheet_name)
    
    # Add QTR Summary to the end if it exists
    if qtr_summary_data:
        filtered_sheets.append(qtr_summary_data)
    
    # Initialize data structure with new column order including Rank
    data = {
        'Station': [],
        'Total Q3 Bookings': [],
        'Rank': [],
        'New Core Billing in Prior Year Same Week': [],
        'Local Q3 Pace': [],
        'National Q3 Pace': [],
        'Digital Q3 Pace': [],
        'Total Q3 Pace': [], 
        'Last Week, Q3 Pace': [],  
        'Pace Delta (bps)': [],  
        'Core Rev Q3 2025': [],  
        'Core Rev Q3 2024': [],  
        'Core Rev Growth %': []  
    }
    
    if last_week_pace_data is None:
        last_week_pace_data = {}
    
    for sheet_name in filtered_sheets:
        try:
            worksheet = current_wb[sheet_name]
            
            # Extract required data
            total_q3_bookings = extract_cell_data(worksheet, 'C40')
            local_pace = extract_cell_data(worksheet, 'K34')
            national_pace = extract_cell_data(worksheet, 'K35')
            digital_pace = extract_cell_data(worksheet, 'K37')
            total_pace = extract_cell_data(worksheet, 'K40')
            
            # Calculate Core Revenue
            d40 = extract_cell_data(worksheet, 'D40')
            d36 = extract_cell_data(worksheet, 'D36')
            core_rev_2025 = d40 - d36
            
            f40 = extract_cell_data(worksheet, 'F40')
            f36 = extract_cell_data(worksheet, 'F36')
            core_rev_2024 = f40 - f36
            
            # Calculate growth percentage
            if core_rev_2024 != 0:
                growth_pct = (core_rev_2025 / core_rev_2024) - 1
            else:
                growth_pct = 0
            
            # Calculate New Core Billing (from Q3 Billings Prior Year logic)
            f40_current = extract_cell_data(worksheet, 'F40')
            f36_current = extract_cell_data(worksheet, 'F36')
            core_current = f40_current - f36_current
            
            # Extract from prior week file (if sheet exists)
            core_prior = 0
            if sheet_name in prior_wb.sheetnames:
                prior_sheet = prior_wb[sheet_name]
                f40_prior = extract_cell_data(prior_sheet, 'F40')
                f36_prior = extract_cell_data(prior_sheet, 'F36')
                core_prior = f40_prior - f36_prior
            else:
                print(f"Warning: Station {sheet_name} not found in prior week file")
            
            new_core_billing = core_current - core_prior
            
            # Get last week's pace data
            last_week_pace = last_week_pace_data.get(sheet_name, 0)
            
            # Calculate Pace Delta in basis points: (Total Q3 Pace - Last Week Q3 Pace) * 100
            pace_delta_bps = (total_pace - last_week_pace) * 100
            
            # Add to data (Rank will be populated after sorting)
            data['Station'].append(sheet_name)
            data['Total Q3 Bookings'].append(total_q3_bookings)
            data['Rank'].append(0)  # Placeholder, will be updated after sorting
            data['New Core Billing in Prior Year Same Week'].append(new_core_billing)
            data['Local Q3 Pace'].append(local_pace)
            data['National Q3 Pace'].append(national_pace)
            data['Digital Q3 Pace'].append(digital_pace)
            data['Total Q3 Pace'].append(total_pace)
            data['Last Week, Q3 Pace'].append(last_week_pace)
            data['Pace Delta (bps)'].append(pace_delta_bps)
            data['Core Rev Q3 2025'].append(core_rev_2025)
            data['Core Rev Q3 2024'].append(core_rev_2024)
            data['Core Rev Growth %'].append(growth_pct)
            
            print(f"Processed station: {sheet_name}")
            
        except Exception as e:
            print(f"Error processing station {sheet_name}: {e}")
            continue
    
    current_wb.close()
    prior_wb.close()
    
    # Create DataFrame from collected data
    df = pd.DataFrame(data)
    
    # Sort by Total Q3 Bookings (highest to lowest) and add ranking
    # Separate QTR Summary row if it exists
    qtr_summary_mask = df['Station'] == 'QTR Summary'
    qtr_summary_row = df[qtr_summary_mask].copy() if qtr_summary_mask.any() else None
    
    # Get all non-QTR Summary rows
    stations_df = df[~qtr_summary_mask].copy()
    
    # Sort stations by Total Q3 Bookings in descending order (highest to lowest)
    stations_df = stations_df.sort_values('Total Q3 Bookings', ascending=False)
    
    # Add ranking (1 to N for stations, excluding QTR Summary)
    stations_df['Rank'] = range(1, len(stations_df) + 1)
    
    # Combine sorted stations with QTR Summary at the bottom
    if qtr_summary_row is not None:
        # Set QTR Summary rank to empty string or 0
        qtr_summary_row['Rank'] = ''
        final_df = pd.concat([stations_df, qtr_summary_row], ignore_index=True)
    else:
        final_df = stations_df
    
    print(f"Sorted {len(stations_df)} stations by Total Q3 Bookings (highest to lowest)")
    
    return final_df

def generate_q3_billings_prior_year(current_file_path, prior_file_path):
    """
    Generate Q3 Billings Prior Year report comparing current week vs prior week.
    """
    print(f"Processing Q3 Billings Prior Year:")
    print(f"  Current week: {current_file_path}")
    print(f"  Prior week: {prior_file_path}")
    
    # Load both workbooks
    current_wb = openpyxl.load_workbook(current_file_path, data_only=True)
    prior_wb = openpyxl.load_workbook(prior_file_path, data_only=True)
    
    # Get station sheets from current file
    station_sheets = get_station_worksheets(current_wb)
    
    # Initialize data structure
    data = {
        'Station': [],
        'Core Bookings Current Week': [],
        'Core Bookings Prior Week': [],
        'New Core Billing in Prior Year Same Week': []
    }
    
    for sheet_name in station_sheets:
        try:
            # Extract from current week file
            current_sheet = current_wb[sheet_name]
            f40_current = extract_cell_data(current_sheet, 'F40')
            f36_current = extract_cell_data(current_sheet, 'F36')
            core_current = f40_current - f36_current
            
            # Extract from prior week file (if sheet exists)
            core_prior = 0
            if sheet_name in prior_wb.sheetnames:
                prior_sheet = prior_wb[sheet_name]
                f40_prior = extract_cell_data(prior_sheet, 'F40')
                f36_prior = extract_cell_data(prior_sheet, 'F36')
                core_prior = f40_prior - f36_prior
            else:
                print(f"Warning: Station {sheet_name} not found in prior week file")
            
            # Calculate difference
            new_billing = core_current - core_prior
            
            # Add to data
            data['Station'].append(sheet_name)
            data['Core Bookings Current Week'].append(core_current)
            data['Core Bookings Prior Week'].append(core_prior)
            data['New Core Billing in Prior Year Same Week'].append(new_billing)
            
            print(f"Processed station: {sheet_name}")
            
        except Exception as e:
            print(f"Error processing station {sheet_name}: {e}")
            continue
    
    current_wb.close()
    prior_wb.close()
    return pd.DataFrame(data)

def extract_date_from_filename(filename):
    """
    Extract date from filename in format MM.DD.YY
    Returns the date string or None if not found
    """
    # Look for pattern like 07.21.25 or 07.14.25
    match = re.search(r'(\d{2}\.\d{2}\.\d{2})', str(filename))
    if match:
        return match.group(1)
    return None

def find_previous_output_file(output_dir, prior_date):
    """
    Find the previous week's output file that ends with the current prior_date.
    For example, if prior_date is 07.14.25, look for Q3_Pacing_Analysis_*_07.14.25.xlsx
    """
    pattern = f'Q3_Pacing_Analysis_*_{prior_date}.xlsx'
    matching_files = list(output_dir.glob(pattern))
    
    if matching_files:
        # Return the most recent matching file
        return matching_files[0]
    return None

def load_last_week_pace_data(previous_output_file):
    """
    Load the 'Total Q3 Pace' column from the previous week's output file.
    Returns a dictionary mapping station names to their Total Q3 Pace values.
    """
    if not previous_output_file or not previous_output_file.exists():
        print("Warning: No previous week output file found")
        return {}
    
    try:
        print(f"Loading last week's pace data from: {previous_output_file}")
        df_previous = pd.read_excel(previous_output_file, sheet_name='Q3 Pacing Setup')
        
        # Create a dictionary mapping station names to Total Q3 Pace values
        pace_data = {}
        for _, row in df_previous.iterrows():
            station = row['Station']
            total_pace = row.get('Total Q3 Pace', 0)
            pace_data[station] = total_pace
        
        print(f"Loaded pace data for {len(pace_data)} stations from previous week")
        return pace_data
        
    except Exception as e:
        print(f"Error loading previous week data: {e}")
        return {}

def create_output_excel(df_pacing, df_billings, output_path):
    """
    Create Excel file with two sheets containing the analysis results.
    """
    print(f"Creating output Excel file: {output_path}")
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write Q3 Pacing Setup sheet
        df_pacing.to_excel(writer, sheet_name='Q3 Pacing Setup', index=False)
        
        # Write Q3 Billings Prior Year sheet
        df_billings.to_excel(writer, sheet_name='Q3 Billings Prior Year', index=False)
        
        # Format the Q3 Pacing Setup sheet
        workbook = writer.book
        pacing_sheet = workbook['Q3 Pacing Setup']
        
        # Define column formatting based on new column structure:
        # A: Station (text)
        # B: Total Q3 Bookings (dollar)
        # C: Rank (whole number)
        # D: New Core Billing in Prior Year Same Week (dollar)
        # E: Local Q3 Pace (percentage)
        # F: National Q3 Pace (percentage)
        # G: Digital Q3 Pace (percentage)
        # H: Total Q3 Pace (percentage)
        # I: Last Week, Q3 Pace (percentage)
        # J: Pace Delta (bps) (whole number)
        # K: Core Rev Q3 2025 (dollar)
        # L: Core Rev Q3 2024 (dollar)
        # M: Core Rev Growth % (percentage)
        
        dollar_columns = ['B', 'D', 'K', 'L']  # Total Q3 Bookings, New Core Billing, Core Rev 2025, Core Rev 2024
        percentage_columns = ['E', 'F', 'G', 'H', 'I', 'M']  # All pace columns and growth %
        whole_number_columns = ['C', 'J']  # Rank, Pace Delta (bps)
        
        # Apply formatting to data rows (skip header row)
        for row in range(2, len(df_pacing) + 2):
            # Format dollar columns: $#,##0
            for col in dollar_columns:
                cell = pacing_sheet[f'{col}{row}']
                cell.number_format = '$#,##0'
            
            # Format percentage columns: 0%
            for col in percentage_columns:
                cell = pacing_sheet[f'{col}{row}']
                cell.number_format = '0%'
            
            # Format whole number columns: #,##0 (for basis points)
            for col in whole_number_columns:
                cell = pacing_sheet[f'{col}{row}']
                cell.number_format = '#,##0'
        
        # Add conditional formatting for Pace Delta (bps) column (Column J)
        from openpyxl.styles import PatternFill, Font
        from openpyxl.formatting.rule import CellIsRule
        
        # Define colors for positive (green) and negative (red) values
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        green_font = Font(color='006100')
        red_font = Font(color='9C0006')
        
        # Apply conditional formatting to Pace Delta (bps) column (Column J)
        pace_delta_range = f'J2:J{len(df_pacing) + 1}'  # J2 to last row with data
        
        # Rule for positive values (green)
        positive_rule = CellIsRule(operator='greaterThan', formula=['0'], 
                                 fill=green_fill, font=green_font)
        pacing_sheet.conditional_formatting.add(pace_delta_range, positive_rule)
        
        # Rule for negative values (red)
        negative_rule = CellIsRule(operator='lessThan', formula=['0'], 
                                 fill=red_fill, font=red_font)
        pacing_sheet.conditional_formatting.add(pace_delta_range, negative_rule)
        
        # Auto-adjust column widths
        for sheet_name in ['Q3 Pacing Setup', 'Q3 Billings Prior Year']:
            sheet = workbook[sheet_name]
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                sheet.column_dimensions[column_letter].width = adjusted_width

def main():
    """
    Main function to orchestrate the pacing analysis.
    """
    print("Starting Q3 Pacing Analysis...")
    
    # Define paths
    inputs_dir = Path('inputs')
    output_dir = Path('output')
    output_dir.mkdir(exist_ok=True)
    
    # Find input files dynamically
    input_files = []
    for file in inputs_dir.glob('*.xlsx'):
        date = extract_date_from_filename(file.name)
        if date:
            input_files.append((file, date))
    
    if len(input_files) < 2:
        print("Error: Need at least 2 Excel files with date patterns in inputs folder")
        return
    
    # Sort files by date to determine which is prior and which is current
    input_files.sort(key=lambda x: x[1])  # Sort by date string
    prior_file, prior_date = input_files[0]  # Earlier date
    current_file, current_date = input_files[-1]  # Later date
    
    print(f"Found input files:")
    print(f"  Prior week: {prior_file} (date: {prior_date})")
    print(f"  Current week: {current_file} (date: {current_date})")
    
    try:
        # Look for previous week's output file to get last week's pace data
        previous_output_file = find_previous_output_file(output_dir, prior_date)
        last_week_pace_data = load_last_week_pace_data(previous_output_file)
        
        # Generate reports
        df_pacing = generate_q3_pacing_setup(current_file, prior_file, last_week_pace_data)
        df_billings = generate_q3_billings_prior_year(current_file, prior_file)
        
        # Create dynamic output filename: Q3_Pacing_Analysis_{prior_date}_{current_date}.xlsx
        output_filename = f'Q3_Pacing_Analysis_{prior_date}_{current_date}.xlsx'
        output_file = output_dir / output_filename
        create_output_excel(df_pacing, df_billings, output_file)
        
        print(f"\nAnalysis complete!")
        print(f"Output saved to: {output_file}")
        print(f"Q3 Pacing Setup: {len(df_pacing)} stations processed")
        print(f"Q3 Billings Prior Year: {len(df_billings)} stations processed")
        
    except Exception as e:
        print(f"Error during analysis: {e}")
        raise

if __name__ == "__main__":
    main()
